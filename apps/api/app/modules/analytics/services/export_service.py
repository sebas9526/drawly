"""Tabular Excel/PDF export. The only two Python dependencies this module adds
(`openpyxl`, `reportlab`) — both are lightweight, pure-Python-friendly, widely
used libraries with no native build step, kept isolated to this one file so
the rest of the module (and the whole backend) has zero export-specific
coupling. Callers pass already-computed rows; this file has no query logic."""

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _cell(value: Any) -> str:
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    return str(value)


def build_excel(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """Real .xlsx bytes — one sheet, bold header row, autosized columns."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None  # a freshly created Workbook() always has one
    sheet.title = title[:31] or "Reporte"  # Excel sheet-name length limit

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(
            [_cell(value) if isinstance(value, datetime | date) else value for value in row]
        )

    for index, _header in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        length = max(
            (len(str(cell.value)) for cell in sheet[column_letter] if cell.value is not None),
            default=8,
        )
        sheet.column_dimensions[column_letter].width = min(max(length + 2, 10), 40)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_pdf(*, title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    """Real .pdf bytes — landscape letter, title + a single styled table."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=16, spaceAfter=12)

    table_data: list[list[str]] = [headers] + [[_cell(value) for value in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F4F8")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    doc.build([Paragraph(title, title_style), Spacer(1, 0.4 * cm), table])
    return buffer.getvalue()
